import openpyxl
from typing import Optional, List


class XlsxTagReader:
    """
    Поиск по тегу в 7-м столбце.
    Возвращает массив [адрес (3-й столбец), тип (4-й столбец)].
    """

    def __init__(self, file_path: str, sheet_name: Optional[str] = None):
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        self.sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active

        # Индекс: {тег: номер_строки} - тег ищется в 7-м столбце
        self.tag_index = {}
        for row in range(1, self.sheet.max_row + 1):
            tag = self.sheet.cell(row=row, column=7).value  # 7-й столбец
            if tag is not None:
                self.tag_index[str(tag).strip()] = row

    def find_by_tag(self, tag: str) -> Optional[List[str]]:
        """
        Ищет строку по тегу в 7-м столбце.

        :param tag: тег для поиска
        :return: массив [адрес, тип] или None
        """
        row = self.tag_index.get(str(tag).strip())
        if row is None:
            return None

        address = self.sheet.cell(row=row, column=3).value  # 3-й столбец - адрес
        var_type = self.sheet.cell(row=row, column=4).value  # 4-й столбец - тип

        return [
            str(address) if address is not None else "",
            str(var_type) if var_type is not None else ""
        ]

    def close(self):
        self.workbook.close()


# Пример
reader = XlsxTagReader("Копия ModbusMapReport (1).xlsx")
result = reader.find_by_tag("P_EUMAX_PROVER_1")

if result:
    print(result)  # ['значение адреса', 'значение типа']
reader.close()